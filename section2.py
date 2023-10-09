'''ファイルの内容をコピペ'''

import openpyxl
from glob import glob
import os

wb = openpyxl.Workbook()
wb.active
ws = wb['Sheet']

header = None

for file_name in glob('01_raw/*'):
    read_wb = openpyxl.load_workbook(file_name)
    read_ws = read_wb.worksheets[0]

    if not header:
        header = read_ws[1]
        header = [v.value for v in header]
        ws.append(header)

    for row in read_ws.iter_rows(min_row=2):
        format_l = []
        for cell in row:
            format_l.append(cell.value)
        ws.append(format_l)

# dir_pathとfile_nameを準備する
dir_path = '02_prep/'
file_name = '期末テスト得点表_まとめ.xlsx'

# フォルダの作成とファイル保存
if not os.path.isdir(dir_path):
    os.makedirs(dir_path)
wb.save(dir_path + file_name)
