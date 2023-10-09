'''コメント一覧を取得'''

import openpyxl

# フォルダとファイルのPATHを記述して、変数に格納する
excel_file_path = 'excel/'
file_name = 'Book1.xlsx'

# 変数read_wbに、読み込んだExcelファイルを格納する
read_wb = openpyxl.load_workbook(excel_file_path + file_name)

# 変数read_wsに、今回使うシートを格納する
read_ws = read_wb.worksheets[0]

# 変数wbを準備して、activeにする
wb = openpyxl.Workbook()
wb.active
ws = wb['Sheet']

header = ['No','セル情報','内容','済']
ws.append(header)

row_number = 1

for row in read_ws.iter_rows(min_row=2,min_col=2):
    for cell in row:
        if cell.comment:
            format_l =[row_number,cell.coordinate,cell.comment.text]
            ws.append(format_l)

            row_number += 1

# 作成したチェックリストの出力
file_name = 'check_list.xlsx'
wb.save(excel_file_path + file_name)